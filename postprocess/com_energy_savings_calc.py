"""
Com-energy-savings-calculation.py

Updates SWHC062_Energy_Savings_Calculations.xlsx with new simulation results.

Inputs (expected in the same directory as this script):
    SWHC062_Energy_Savings_Calculations_20260818.xlsx  - workbook to update
    Summary-Report-all.csv          - Base/Measure run results, all 17 building
                                       types (columns A-O: Filename, Measure,
                                       BldgType, BldgLoc, BldgHVAC, Run Type,
                                       Cooling Capacity, Heating Capacity,
                                       Conditioned Area, Electricity Heating,
                                       Natural Gas Heating, Electricity Cooling,
                                       Electricity Fans, Unmet Hours Heating,
                                       Unmet Hours Cooling). Has one metadata
                                       row above the real header, so it is read
                                       with skiprows=1.
    Deer_Peak_-_Electric-all.csv    - DEER peak-period electric results, all
                                       17 building types (columns A-G: Building
                                       Type, Measure, System Type, Run Type,
                                       Climate Zone, Average Temperature,
                                       Average Electric Energy).

What this script does:
    - Base tab, columns A-O:      Summary-Report-all.csv rows where
                                   Run Type == "Base"
    - Measure tab, columns A-O:   Summary-Report-all.csv rows where
                                   Run Type == "Measure"
    - Pk Base tab, columns A-G:   Deer_Peak_-_Electric-all.csv rows where
                                   Run Type == "Base"
    - Pk Measure tab, columns A-G: Deer_Peak_-_Electric-all.csv rows where
                                   Run Type == "Measure"
    - Readme tab: appends a description of what was added, below row 16.
    - "Measure Off-hour" tab and all calculation tabs (Lookup, UEC_*, etc.)
      are left untouched.

Row order: both source files are sorted to match the workbook's existing
convention -- Building Type (in the specific 17-type order already used in
the "Measure Off-hour" tab) -> Measure (M1-M5) -> BldgHVAC/System Type
(cDXGF, cDXHP, cPTAC, cPVVG) -> Climate Zone (CZ01-CZ16).

Both source files include a real "Htl" (Hotel) building type, and an
"MFmCmn" building type whose values are a direct copy of "OfS" (confirmed
identical row-for-row) -- this was provided that way by the data source,
not synthesized by this script.

After running this script, recalculate the workbook (e.g. with
scripts/recalc.py from the xlsx skill) so formula-driven tabs pick up the
new inputs -- this script only writes columns A-O / A-G of the four target
tabs; it does not touch or recompute any formulas.
"""

import openpyxl
import pandas as pd

WORKBOOK_IN = "SWHC062_Energy_Savings_Calculations_20260818.xlsx"
WORKBOOK_OUT = "SWHC062_Energy_Savings_Calculations_20260827_updated.xlsx"
SUMMARY_CSV = "Summary-Report-all.csv"
PEAK_CSV = "Deer_Peak_-_Electric-all.csv"

BLDG_TYPE_ORDER = [
    "Asm", "ECC", "Epr", "ERC", "Ese", "Eun", "Gro", "Htl", "MBT",
    "MFmCmn", "MLI", "OfS", "RFF", "Rt3", "RtL", "RtS", "SCn",
]
MEASURE_ORDER = ["M1", "M2", "M3", "M4", "M5"]
HVAC_ORDER = ["cDXGF", "cDXHP", "cPTAC", "cPVVG"]
CZ_ORDER = [f"CZ{n:02d}" for n in range(1, 17)]

SUMMARY_COLS = [
    "Filename", "Measure", "BldgType", "BldgLoc", "BldgHVAC", "Run Type",
    "Cooling Capacity", "Heating Capacity", "Conditioned Area",
    "Electricity Heating", "Natural Gas Heating", "Electricity Cooling",
    "Electricity Fans", "Unmet Hours Heating", "Unmet Hours Cooling",
]
PEAK_COLS = [
    "Building Type", "Measure", "System Type", "Run Type", "Climate Zone",
    "Average Temperature", "Average Electric Energy",
]

README_ADDITIONS = [
    ("Pk Base", "Baseline whole building average peak demand during DEER peak periods "
                 "(source: Deer_Peak_-_Electric-all.csv)"),
    ("Pk Measure", "Measure case whole building average peak demand during DEER peak "
                    "periods (source: Deer_Peak_-_Electric-all.csv)"),
    ("Base / Measure", "Updated 2026-08-20 with full 17-building-type results from "
                        "Summary-Report-all.csv (adds real 'Htl' data and an "
                        "'MFmCmn' building type, whose values are a direct copy of "
                        "'OfS' as provided by the data source)."),
]


def sort_summary(df):
    df = df.copy()
    df["BldgType"] = pd.Categorical(df["BldgType"], categories=BLDG_TYPE_ORDER, ordered=True)
    df["Measure"] = pd.Categorical(df["Measure"], categories=MEASURE_ORDER, ordered=True)
    df["BldgHVAC"] = pd.Categorical(df["BldgHVAC"], categories=HVAC_ORDER, ordered=True)
    df["BldgLoc"] = pd.Categorical(df["BldgLoc"], categories=CZ_ORDER, ordered=True)
    return df.sort_values(["BldgType", "Measure", "BldgHVAC", "BldgLoc"]).reset_index(drop=True)


def sort_peak(df):
    df = df.copy()
    df["Building Type"] = pd.Categorical(df["Building Type"], categories=BLDG_TYPE_ORDER, ordered=True)
    df["Measure"] = pd.Categorical(df["Measure"], categories=MEASURE_ORDER, ordered=True)
    df["System Type"] = pd.Categorical(df["System Type"], categories=HVAC_ORDER, ordered=True)
    df["Climate Zone"] = pd.Categorical(df["Climate Zone"], categories=CZ_ORDER, ordered=True)
    return df.sort_values(["Building Type", "Measure", "System Type", "Climate Zone"]).reset_index(drop=True)


def write_block(ws, df, cols, start_row=4):
    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + i
        for c, col_name in enumerate(cols, start=1):
            ws.cell(row=r, column=c, value=row[col_name])


def main():
    summary = pd.read_csv(SUMMARY_CSV, skiprows=1)
    peak = pd.read_csv(PEAK_CSV)

    base_df = sort_summary(summary[summary["Run Type"] == "Base"])
    meas_df = sort_summary(summary[summary["Run Type"] == "Measure"])
    pk_base_df = sort_peak(peak[peak["Run Type"] == "Base"])
    pk_meas_df = sort_peak(peak[peak["Run Type"] == "Measure"])

    print(f"Base rows: {len(base_df)}, Measure rows: {len(meas_df)}, "
          f"Pk Base rows: {len(pk_base_df)}, Pk Measure rows: {len(pk_meas_df)}")

    wb = openpyxl.load_workbook(WORKBOOK_IN, data_only=False)

    write_block(wb["Base"], base_df, SUMMARY_COLS)
    write_block(wb["Measure"], meas_df, SUMMARY_COLS)
    write_block(wb["Pk Base"], pk_base_df, PEAK_COLS)
    write_block(wb["Pk Measure"], pk_meas_df, PEAK_COLS)

    # Readme: append description of what was added, below row 16
    ws = wb["Readme"]
    r = 17
    for tab, description in README_ADDITIONS:
        ws.cell(row=r, column=1, value=tab)
        ws.cell(row=r, column=2, value=description)
        r += 1

    wb.save(WORKBOOK_OUT)
    print(f"Saved {WORKBOOK_OUT}")
    print("Remember to recalculate the workbook (e.g. scripts/recalc.py) before use.")


if __name__ == "__main__":
    main()
